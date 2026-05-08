#!/usr/bin/env python3
"""XLSX → PDF 转换脚本（LibreOffice 不可用时的 fallback）
使用 openpyxl 读取 Excel，reportlab 生成 PDF。
保持与原 xlsx 基本一致的行列布局，支持中文字体。

用法: python3 xlsx2pdf.py <input.xlsx> <output_dir>
"""

import sys
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── 字体注册 ──
FONT_PATHS = [
    "/System/Library/Fonts/STHeiti Light.ttc",       # macOS 黑体
    "/System/Library/Fonts/PingFang.ttc",             # macOS 苹方
    "/System/Library/Fonts/Supplemental/Songti.ttc",  # macOS 宋体
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
]

FONT_NAME = None
FONT_BOLD = None

for fp in FONT_PATHS:
    if os.path.exists(fp):
        try:
            pdfmetrics.registerFont(TTFont("ChineseFont", fp))
            FONT_NAME = "ChineseFont"
            break
        except Exception:
            continue

if FONT_NAME is None:
    print("警告: 未找到中文字体，PDF中中文将显示为空白", file=sys.stderr)
    FONT_NAME = "Helvetica"


def read_xlsx(xlsx_path):
    """读取 xlsx 所有 sheet，返回 {sheet_name: (headers, data_rows)}"""
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        if rows:
            sheets[sn] = rows
    wb.close()
    return sheets


def build_pdf(xlsx_path, output_pdf, sheets):
    """用 reportlab 生成 PDF"""
    # 横向 A4 以容纳更多列
    page_w, page_h = landscape(A4)
    margin = 10 * mm

    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )

    elements = []
    style_sheet = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=style_sheet["Heading2"],
        fontName=FONT_NAME,
        fontSize=10,
        leading=14,
        spaceAfter=6,
    )

    cell_style = ParagraphStyle(
        "ChineseCell",
        fontName=FONT_NAME,
        fontSize=7,
        leading=9,
        wordWrap="CJK",
    )

    for sn, rows in sheets.items():
        if not rows:
            continue

        # Sheet 标题
        elements.append(Paragraph(f"Sheet: {sn}", title_style))
        elements.append(Spacer(1, 3 * mm))

        # 计算列数
        max_cols = max(len(r) for r in rows)
        # 补齐每行列数一致
        padded = []
        for r in rows:
            while len(r) < max_cols:
                r.append("")
            padded.append(r)

        # 第一行作为表头
        header = [Paragraph(c, cell_style) for c in padded[0]]
        data = []
        for r in padded[1:]:
            data.append([Paragraph(c, cell_style) for c in r])

        table_data = [header] + data

        # 动态列宽
        base_w = float(page_w - 2 * margin)
        col_widths = [base_w / max_cols] * max_cols
        # 首列稍宽
        if max_cols > 1:
            col_widths[0] = col_widths[0] * 1.3
            rest_w = (page_w - 2 * margin - col_widths[0]) / (max_cols - 1)
            for i in range(1, max_cols):
                col_widths[i] = rest_w

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # 样式
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)
        elements.append(Spacer(1, 10 * mm))

    doc.build(elements)
    return output_pdf


def main():
    if len(sys.argv) < 3:
        print(f"用法: {sys.argv[0]} <input.xlsx> <output_dir>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.exists(xlsx_path):
        print(f"错误: 文件不存在 {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_pdf = os.path.join(output_dir, f"{base}.pdf")

    sheets = read_xlsx(xlsx_path)
    build_pdf(xlsx_path, output_pdf, sheets)

    print(output_pdf)


if __name__ == "__main__":
    main()
